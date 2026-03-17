"""
Report Export Module
=====================
Export analytics reports in PDF and Excel formats for monetization.

Usage:
    from report_exporter import generate_excel_report, generate_pdf_report
    
    # Excel report
    generate_excel_report("revenue_report.xlsx")
    
    # PDF report
    generate_pdf_report("revenue_report.pdf", title="Monthly Revenue Report")
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Optional, Dict, Any, List
import os
import tempfile
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Designated export directory - set via environment variable or use user's documents folder
# In production, set REPORT_EXPORT_DIR to a writable directory
DEFAULT_EXPORT_DIR = os.path.join(tempfile.gettempdir(), 'revenue_reports')
EXPORT_DIR = os.environ.get("REPORT_EXPORT_DIR", DEFAULT_EXPORT_DIR)

# Ensure default export directory exists
if not os.path.isdir(EXPORT_DIR):
    try:
        os.makedirs(EXPORT_DIR, exist_ok=True)
    except OSError:
        # Fallback to temp directory if default creation fails
        EXPORT_DIR = tempfile.gettempdir()
        logger.warning(f"Could not create default export directory. Using: {EXPORT_DIR}")

# Track if validation has been performed
_export_dir_validated = False

def _validate_export_dir():
    """Validate that the export directory exists and is writable.
    
    This validation is performed lazily on first export attempt rather than at import time.
    """
    global _export_dir_validated
    
    if _export_dir_validated:
        return
        
    if not os.path.isdir(EXPORT_DIR):
        raise RuntimeError(
            f"Export directory does not exist: {EXPORT_DIR}. "
            f"Set REPORT_EXPORT_DIR environment variable to a valid directory."
        )
    if not os.access(EXPORT_DIR, os.W_OK):
        raise RuntimeError(
            f"Export directory is not writable: {EXPORT_DIR}. "
            f"Check directory permissions."
        )
    
    _export_dir_validated = True

try:
    from openpyxl import Workbook as OpenpyxlWB
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from jinja2 import Template
    from weasyprint import HTML
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


# ==================== Excel Export ====================

class ExcelReportGenerator:
    """Generate professional Excel reports."""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.workbook = None
    
    def generate_revenue_report(self, data: Dict[str, Any]) -> str:
        """Generate comprehensive revenue report."""
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel export not available. Install openpyxl and xlsxwriter.")
            
        workbook = OpenpyxlWB()
        self._create_formats_openpyxl(workbook)
        
        # Summary Sheet
        self._add_summary_sheet(workbook, data)
        
        # Revenue Details Sheet
        if 'monthly_revenue' in data:
            self._add_monthly_revenue_sheet(workbook, data['monthly_revenue'])
            
        # Customer Segments Sheet
        if 'segments' in data:
            self._add_segments_sheet(workbook, data['segments'])
            
        # Churn Analysis Sheet
        if 'churn' in data:
            self._add_churn_sheet(workbook, data['churn'])
            
        # Save
        workbook.save(self.filepath)
        return self.filepath
        
    def _create_formats_openpyxl(self, workbook):
        """Create OpenPyXL formatting styles."""
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=16, color="1F4E79")
        self.subtitle_font = Font(bold=True, size=12, color="2E75B6")
        self.currency_format = '$#,##0.00'
        self.percent_format = '0.0%'
        
    def _add_summary_sheet(self, workbook, data: Dict):
        """Add executive summary sheet."""
        ws = workbook.active
        ws.title = "Executive Summary"
        
        # Title
        ws['A1'] = "Revenue Intelligence Report"
        ws['A1'].font = self.title_font
        
        # Date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        # Key Metrics
        row = 4
        ws[f'A{row}'] = "Key Performance Indicators"
        ws[f'A{row}'].font = self.subtitle_font
        
        metrics = [
            ("Total Revenue", data.get('total_revenue', 0)),
            ("Total Customers", data.get('total_customers', 0)),
            ("Average Order Value", data.get('avg_order_value', 0)),
            ("Churn Rate", data.get('churn_rate', 0)),
        ]
        
        row += 1
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            if 'Revenue' in metric or 'Value' in metric:
                ws[f'B{row}'].number_format = self.currency_format
            row += 1
            
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
    def _add_monthly_revenue_sheet(self, workbook, monthly_data):
        """Add monthly revenue sheet."""
        ws = workbook.create_sheet("Monthly Revenue")
        
        # Headers
        headers = ['Month', 'Revenue', 'Orders', 'Growth %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            
        # Data
        for row, record in enumerate(monthly_data, 2):
            ws.cell(row=row, column=1, value=record.get('month', ''))
            ws.cell(row=row, column=2, value=record.get('revenue', 0))
            ws.cell(row=row, column=2).number_format = self.currency_format
            ws.cell(row=row, column=3, value=record.get('orders', 0))
            ws.cell(row=row, column=4, value=record.get('growth', 0))
            ws.cell(row=row, column=4).number_format = self.percent_format
            
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        
    def _add_segments_sheet(self, workbook, segments: Dict):
        """Add customer segments sheet."""
        ws = workbook.create_sheet("Customer Segments")
        
        # Headers
        headers = ['Segment', 'Customer Count', 'Total Revenue', 'Avg Revenue', 'Avg Frequency']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            
        # Data
        row = 2
        for segment, metrics in segments.items():
            ws.cell(row=row, column=1, value=segment)
            ws.cell(row=row, column=2, value=metrics.get('count', 0))
            ws.cell(row=row, column=3, value=metrics.get('revenue', 0))
            ws.cell(row=row, column=3).number_format = self.currency_format
            ws.cell(row=row, column=4, value=metrics.get('avg_revenue', 0))
            ws.cell(row=row, column=4).number_format = self.currency_format
            ws.cell(row=row, column=5, value=metrics.get('avg_frequency', 0))
            row += 1
            
    def _add_churn_sheet(self, workbook, churn_data: Dict):
        """Add churn analysis sheet."""
        ws = workbook.create_sheet("Churn Analysis")
        
        # Headers
        headers = ['Customer ID', 'Churn Risk Score', 'Segment', 'Recommended Action']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            
        # Data
        row = 2
        for customer in churn_data.get('high_risk', [])[:1000]:  # Limit to 1000
            ws.cell(row=row, column=1, value=customer.get('customer_id', ''))
            ws.cell(row=row, column=2, value=customer.get('churn_probability', 0))
            ws.cell(row=row, column=2).number_format = self.percent_format
            ws.cell(row=row, column=3, value=customer.get('segment', ''))
            ws.cell(row=row, column=4, value=customer.get('action', ''))
            row += 1


def _load_data_for_reports():
    """Load data for report generation without importing api_service."""
    import pandas as pd
    data = {}
    # Use absolute path relative to this module to avoid CWD issues
    module_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(module_dir, 'Data')
    
    try:
        data['customer_rfm'] = pd.read_csv(f'{data_dir}/customer_rfm.csv')
        data['rfm_segments'] = pd.read_csv(f'{data_dir}/rfm_segments.csv')
        data['monthly_revenue'] = pd.read_csv(f'{data_dir}/monthly_revenue.csv')
        data['churn_summary'] = pd.read_csv(f'{data_dir}/churn_summary.csv')
        data['cohort_retention'] = pd.read_csv(f'{data_dir}/cohort_retention_with_churn.csv')
        data['top_categories'] = pd.read_csv(f'{data_dir}/top_categories.csv')
        data['top_states'] = pd.read_csv(f'{data_dir}/top_states.csv')
        
        try:
            data['customer_churn'] = pd.read_csv(f'{data_dir}/customer_churn.csv')
        except FileNotFoundError:
            data['customer_churn'] = None
        except pd.errors.EmptyDataError:
            data['customer_churn'] = None
        except Exception as e:
            # Log the actual error for debugging
            import logging
            logging.warning(f"Failed to load customer_churn.csv: {e}")
            data['customer_churn'] = None
            
        return data
    except Exception as e:
        raise ImportError(f"Error loading data files: {e}")


def generate_excel_report(filename: str, data: Optional[Dict] = None) -> str:
    """
    Generate Excel report file.
    
    Args:
        filename: Output filename
        data: Report data dictionary
        
    Returns:
        Generated filename (full path)
    """
    import re  # Import regex for filename sanitization
    
    # Sanitize filename to prevent path traversal
    # Get basename to remove any directory components
    safe_filename = os.path.basename(filename)
    
    # Security check: if directory separators survived basename extraction,
    # it might be a path traversal attempt
    if safe_filename != filename and not filename.startswith(safe_filename):
        safe_filename = "report.xlsx"
    
    # Filter to only allow safe characters
    safe_filename = re.sub(r'[^\w\-.]', '', safe_filename)
    if not safe_filename:
        safe_filename = "report.xlsx"
    
    # Ensure .xlsx extension
    if not safe_filename.endswith('.xlsx'):
        safe_filename += '.xlsx'
    
    # Validate export directory lazily on first use
    _validate_export_dir()
    
    # Construct full path in designated export directory
    output_path = os.path.join(EXPORT_DIR, safe_filename)
    if data is None:
        # Load default data using local function to avoid circular imports
        df = _load_data_for_reports()
        
        # Transform for report
        data = {}
        data['total_revenue'] = df['monthly_revenue']['revenue'].sum()
        data['total_customers'] = len(df['rfm_segments'])
        data['avg_order_value'] = data['total_revenue'] / data['total_customers'] if data['total_customers'] > 0 else 0
        
        # Calculate churn rate dynamically from data
        if 'customer_churn' in df and df['customer_churn'] is not None and len(df['customer_churn']) > 0:
            # Use actual churn data if available
            churn_df = df['customer_churn']
            if 'churn_probability' in churn_df.columns:
                # Use probability threshold of 0.6 for high risk
                data['churn_rate'] = (churn_df['churn_probability'] >= 0.6).mean()
            elif 'churn_label' in churn_df.columns:
                data['churn_rate'] = churn_df['churn_label'].mean()
            else:
                data['churn_rate'] = 0.0
        elif 'rfm_segments' in df:
            # Calculate from RFM segments as fallback
            segment_risk = {
                'Champions': 0.05,
                'Loyal Customers': 0.10,
                'Potential Loyalist': 0.20,
                'New Customers': 0.30,
                'At Risk': 0.60,
                'Cant Lose Them': 0.80,
                'Lost': 0.95
            }
            churn_risks = df['rfm_segments']['rfm_segment'].map(segment_risk).fillna(0.5)
            data['churn_rate'] = (churn_risks >= 0.6).mean()
        else:
            data['churn_rate'] = 0.0
        
    generator = ExcelReportGenerator(output_path)
    return generator.generate_revenue_report(data)


# ==================== PDF Export ====================

class PDFReportGenerator:
    """Generate professional PDF reports."""
    
    HTML_TEMPLATE = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>{{ title }}</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 40px;
                color: #333;
            }
            .header {
                text-align: center;
                margin-bottom: 30px;
                border-bottom: 2px solid #1F4E79;
                padding-bottom: 20px;
            }
            .header h1 {
                color: #1F4E79;
                margin: 0;
            }
            .header .subtitle {
                color: #666;
                font-size: 14px;
                margin-top: 5px;
            }
            .section {
                margin-bottom: 25px;
            }
            .section h2 {
                color: #2E75B6;
                font-size: 16px;
                border-left: 4px solid #2E75B6;
                padding-left: 10px;
                margin-bottom: 15px;
            }
            .metric-grid {
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 15px;
                margin-bottom: 20px;
            }
            .metric-box {
                background: #f5f5f5;
                padding: 15px;
                border-radius: 5px;
                text-align: center;
            }
            .metric-box .label {
                font-size: 12px;
                color: #666;
                margin-bottom: 5px;
            }
            .metric-box .value {
                font-size: 20px;
                font-weight: bold;
                color: #1F4E79;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
            }
            th {
                background: #1F4E79;
                color: white;
                padding: 10px;
                text-align: left;
                font-size: 12px;
            }
            td {
                padding: 8px 10px;
                border-bottom: 1px solid #ddd;
                font-size: 11px;
            }
            tr:nth-child(even) {
                background: #f9f9f9;
            }
            .chart-placeholder {
                background: #f0f0f0;
                height: 200px;
                display: flex;
                align-items: center;
                justify-content: center;
                color: #999;
                margin-bottom: 20px;
            }
            .footer {
                margin-top: 40px;
                padding-top: 20px;
                border-top: 1px solid #ddd;
                text-align: center;
                font-size: 10px;
                color: #999;
            }
            .insight-box {
                background: #e8f4f8;
                padding: 15px;
                border-radius: 5px;
                margin-bottom: 15px;
            }
            .insight-box h3 {
                margin: 0 0 10px 0;
                color: #1F4E79;
                font-size: 14px;
            }
            .insight-box ul {
                margin: 0;
                padding-left: 20px;
            }
            .insight-box li {
                margin-bottom: 5px;
            }
        </style>
    </head>
    <body>
        <div class="header">
            <h1>{{ title }}</h1>
            <div class="subtitle">{{ subtitle }}</div>
        </div>
        
        <div class="section">
            <h2>Key Performance Indicators</h2>
            <div class="metric-grid">
                <div class="metric-box">
                    <div class="label">Total Revenue</div>
                    <div class="value">{{ total_revenue }}</div>
                </div>
                <div class="metric-box">
                    <div class="label">Total Customers</div>
                    <div class="value">{{ total_customers }}</div>
                </div>
                <div class="metric-box">
                    <div class="label">Avg Order Value</div>
                    <div class="value">{{ avg_order_value }}</div>
                </div>
                <div class="metric-box">
                    <div class="label">Churn Rate</div>
                    <div class="value">{{ churn_rate }}</div>
                </div>
            </div>
        </div>
        
        {% if segments %}
        <div class="section">
            <h2>Customer Segment Analysis</h2>
            <table>
                <tr>
                    <th>Segment</th>
                    <th>Customers</th>
                    <th>Revenue</th>
                    <th>Avg Value</th>
                </tr>
                {% for segment in segments %}
                <tr>
                    <td>{{ segment.name }}</td>
                    <td>{{ segment.count }}</td>
                    <td>{{ segment.revenue }}</td>
                    <td>{{ segment.avg }}</td>
                </tr>
                {% endfor %}
            </table>
        </div>
        {% endif %}
        
        {% if insights %}
        <div class="section">
            <h2>Business Insights & Recommendations</h2>
            {% for insight in insights %}
            <div class="insight-box">
                <h3>{{ insight.title }}</h3>
                <ul>
                    {% for item in insight.items %}
                    <li>{{ item }}</li>
                    {% endfor %}
                </ul>
            </div>
            {% endfor %}
        </div>
        {% endif %}
        
        {% if monthly_revenue %}
        <div class="section">
            <h2>Monthly Revenue Trend</h2>
            <table>
                <tr>
                    <th>Month</th>
                    <th>Revenue</th>
                    <th>Growth</th>
                </tr>
                {% for month in monthly_revenue %}
                <tr>
                    <td>{{ month.date }}</td>
                    <td>{{ month.revenue }}</td>
                    <td>{{ month.growth }}</td>
                </tr>
                {% endfor %}
            </table>
        </div>
        {% endif %}
        
        <div class="footer">
            Generated by Revenue Intelligence System | {{ date }}
        </div>
    </body>
    </html>
    """
    
    def __init__(self, filename: str):
        self.filename = filename
        
    def generate(self, title: str, data: Dict[str, Any]) -> str:
        """Generate PDF report."""
        if not PDF_AVAILABLE:
            raise ImportError("PDF export not available. Install jinja2 and weasyprint.")
            
        # Prepare data
        template = Template(self.HTML_TEMPLATE)
        
        # Format data for template
        render_data = {
            'title': title,
            'subtitle': data.get('subtitle', 'Executive Report'),
            'total_revenue': f"${data.get('total_revenue', 0):,.2f}",
            'total_customers': f"{data.get('total_customers', 0):,}",
            'avg_order_value': f"${data.get('avg_order_value', 0):,.2f}",
            'churn_rate': f"{data.get('churn_rate', 0)*100:.1f}%",
            'date': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'segments': data.get('segments', []),
            'monthly_revenue': data.get('monthly_revenue', []),
            'insights': data.get('insights', [])
        }
        
        # Render HTML
        html_content = template.render(**render_data)
        
        # Convert to PDF
        HTML(string=html_content).write_pdf(self.filename)
        
        return self.filename


def generate_pdf_report(filename: str, title: str = "Revenue Intelligence Report", 
                         data: Optional[Dict] = None) -> str:
    """
    Generate PDF report file.
    
    Args:
        filename: Output filename
        title: Report title
        data: Report data dictionary
        
    Returns:
        Generated filename (full path)
    """
    import re  # Import regex for filename sanitization
    
    # Sanitize filename to prevent path traversal
    # Get basename to remove any directory components
    safe_filename = os.path.basename(filename)
    
    # Security check: if directory separators survived basename extraction,
    # it might be a path traversal attempt
    if safe_filename != filename and not filename.startswith(safe_filename):
        safe_filename = "report.pdf"
    
    # Filter to only allow safe characters
    safe_filename = re.sub(r'[^\w\-.]', '', safe_filename)
    if not safe_filename:
        safe_filename = "report.pdf"
    
    # Ensure .pdf extension
    if not safe_filename.endswith('.pdf'):
        safe_filename += '.pdf'
    
    # Validate export directory lazily on first use
    _validate_export_dir()
    
    # Construct full path in designated export directory
    output_path = os.path.join(EXPORT_DIR, safe_filename)
    if data is None:
        # Load default data using local function to avoid circular imports
        df = _load_data_for_reports()
        
        # Calculate churn rate dynamically from data
        churn_rate = 0.0
        if 'customer_churn' in df and df['customer_churn'] is not None and len(df['customer_churn']) > 0:
            # Use actual churn data if available
            churn_df = df['customer_churn']
            if 'churn_probability' in churn_df.columns:
                churn_rate = (churn_df['churn_probability'] >= 0.6).mean()
            elif 'churn_label' in churn_df.columns:
                churn_rate = churn_df['churn_label'].mean()
        elif 'rfm_segments' in df:
            # Calculate from RFM segments as fallback
            segment_risk = {
                'Champions': 0.05,
                'Loyal Customers': 0.10,
                'Potential Loyalist': 0.20,
                'New Customers': 0.30,
                'At Risk': 0.60,
                'Cant Lose Them': 0.80,
                'Lost': 0.95
            }
            churn_risks = df['rfm_segments']['rfm_segment'].map(segment_risk).fillna(0.5)
            churn_rate = (churn_risks >= 0.6).mean()
        
        # Transform for report
        data = {
            'subtitle': 'Monthly Performance Report',
            'total_revenue': df['monthly_revenue']['revenue'].sum(),
            'total_customers': len(df['rfm_segments']),
            'avg_order_value': df['monthly_revenue']['revenue'].sum() / len(df['rfm_segments']) if len(df['rfm_segments']) > 0 else 0,
            'churn_rate': churn_rate,
            'segments': _get_segments_for_pdf(df),
            'monthly_revenue': _get_monthly_for_pdf(df),
            'insights': _get_default_insights()
        }
        
    generator = PDFReportGenerator(output_path)
    return generator.generate(title, data)


def _get_segments_for_pdf(df: Dict) -> List[Dict]:
    """Get segments formatted for PDF."""
    segments = df['rfm_segments'].groupby('rfm_segment').agg({
        'monetary': ['count', 'sum', 'mean']
    })
    segments.columns = ['count', 'revenue', 'avg']
    segments = segments.reset_index()
    
    return [
        {
            'name': row['rfm_segment'],
            'count': int(row['count']),
            'revenue': f"${row['revenue']:,.2f}",
            'avg': f"${row['avg']:,.2f}"
        }
        for _, row in segments.iterrows()
    ]


def _get_monthly_for_pdf(df: Dict) -> List[Dict]:
    """Get monthly revenue formatted for PDF."""
    monthly = df['monthly_revenue'].copy()
    monthly = monthly.sort_values('month')
    
    result = []
    prev = None
    for _, row in monthly.iterrows():
        growth = ''
        if prev and prev > 0:
            growth = f"{((row['revenue'] - prev) / prev * 100):+.1f}%"
        result.append({
            'date': row['month'],
            'revenue': f"${row['revenue']:,.2f}",
            'growth': growth
        })
        prev = row['revenue']
        
    return result


def _get_default_insights() -> List[Dict]:
    """Get default business insights."""
    return [
        {
            'title': 'Retention Strategy',
            'items': [
                'Focus retention efforts on At Risk and Cant Lose Them segments',
                'High-spending customers (monetary feature) show distinct churn patterns',
                'Payment installments indicate price sensitivity - offer flexible payment options'
            ]
        },
        {
            'title': 'Revenue Optimization',
            'items': [
                'Champions and Loyal Customers drive majority of revenue',
                'Cross-sell premium products to Potential Loyalists',
                'New Customers need onboarding sequences for long-term retention'
            ]
        },
        {
            'title': 'Customer Experience',
            'items': [
                'Review scores correlate with retention - monitor satisfaction closely',
                'Late deliveries marginally increase churn - improve logistics',
                'Credit card users may show different loyalty patterns'
            ]
        }
    ]


# ==================== Quick Export Functions ====================

def export_revenue_report(format: str = 'excel', filename: Optional[str] = None) -> str:
    """
    Quick export function for revenue report.
    
    Args:
        format: 'excel' or 'pdf'
        filename: Output filename (auto-generated if None)
        
    Returns:
        Generated filename
    """
    if filename is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"revenue_report_{timestamp}.{format}"
        
    if format.lower() == 'excel':
        return generate_excel_report(filename)
    elif format.lower() == 'pdf':
        return generate_pdf_report(filename)
    else:
        raise ValueError(f"Unsupported format: {format}")


# Run standalone for testing
if __name__ == "__main__":
    print("Generating sample reports...")
    
    try:
        # Generate Excel
        excel_file = export_revenue_report('excel')
        print(f"✅ Excel report: {excel_file}")
    except Exception as e:
        print(f"⚠️ Excel export skipped: {e}")
        
    try:
        # Generate PDF
        pdf_file = export_revenue_report('pdf')
        print(f"✅ PDF report: {pdf_file}")
    except Exception as e:
        print(f"⚠️ PDF export skipped: {e}")
        
    print("Done!")
